import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog


def upload_excel_file():
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    return file_path


# Upload the first Excel file
print("Upload the first Excel file:")
file1_path = upload_excel_file()


# Upload the second Excel file
print("Upload the second Excel file:")
file2_path = upload_excel_file()

# Read the first Excel file and reformat column names to uppercase
df1 = pd.read_excel(file1_path)
df1.columns = df1.columns.str.upper()

# Read the second Excel file and reformat column names to uppercase
df2 = pd.read_excel(file2_path)
df2.columns = df2.columns.str.upper()


# Function to update column names
def update_column_names(df):
    new_columns = []
    for column in df.columns:
        if column.startswith("ID"):
            new_column = "@" + column
        else:
            new_column = column
        new_columns.append(new_column)
    df.columns = new_columns


# Update column names in the first Excel file
update_column_names(df1)

# Update column names in the second Excel file
update_column_names(df2)

# Prepend the path before every ".psd" value in the Excel files
path = "Macintosh HD 2:Users:rickschurink:Desktop:Links:"
id_columns = [column for column in df1.columns if "@ID" in column]

for column in id_columns:
    df1[column] = df1[column].apply(lambda x: x.replace(path, "") if isinstance(x, str) else x)
    df1[column] = df1[column].apply(lambda x: f"{path}{x}" if isinstance(x, str) and ".psd" in x else x)

    df2[column] = df2[column].apply(lambda x: x.replace(path, "") if isinstance(x, str) else x)
    df2[column] = df2[column].apply(lambda x: f"{path}{x}" if isinstance(x, str) and ".psd" in x else x)


# Combine the column names from both dataframes
combined_columns = list(set(df1.columns) | set(df2.columns))

# Convert column names to start with a capital letter and the rest lowercase
combined_columns = [column.capitalize() for column in combined_columns]

# Drop duplicates from key columns in both dataframes
key_columns = ['ACTIE', 'MERK', 'VAN', 'VOOR', 'UITGELICHT', 'PRODUCT AANDUIDING']
if 'PRODUCT AANDUIDING' in df1.columns:
    df1.drop_duplicates(subset=key_columns, inplace=True)
if 'PRODUCT AANDUIDING' in df2.columns:
    df2.drop_duplicates(subset=key_columns, inplace=True)

# Get rows present in df1 but not in df2
missing_rows = pd.DataFrame(columns=combined_columns)
for column in key_columns:
    if column in df1.columns and column in df2.columns:
        missing_rows = pd.concat([missing_rows, df1[~df1[column].isin(df2[column])]])

# Merge the two dataframes
merged_df = pd.concat([df1, df2])

# Remove duplicates
merged_df.drop_duplicates(inplace=True)

# Create a new Excel workbook
wb = Workbook()
ws = wb.active

# Calculate maximum content length for each column
column_lengths = [0] * len(merged_df.columns)
for col_idx, column_name in enumerate(merged_df.columns):
    max_length = max(
        merged_df[column_name].astype(str).apply(len).max(),
        len(column_name)
    )
    column_lengths[col_idx] = max_length

# Set column widths based on maximum content length
for col_idx, length in enumerate(column_lengths):
    column_letter = get_column_letter(col_idx + 1)
    column_width = (length + 2) * 1.2  # Adjust the factor as needed
    ws.column_dimensions[column_letter].width = column_width

# Apply formatting to header row
header_font = Font(bold=True)
header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
for col_idx, column_name in enumerate(merged_df.columns, 1):
    cell = ws.cell(row=1, column=col_idx, value=column_name)
    cell.font = header_font
    cell.fill = header_fill

# Apply alternating row colors
odd_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
even_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
for row_idx in range(2, len(merged_df) + 2):
    for col_idx, cell_value in enumerate(merged_df.iloc[row_idx - 2], 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
        if row_idx % 2 == 0:  # Even row number
            cell.fill = even_fill
        else:  # Odd row number
            cell.fill = odd_fill

# Apply formatting to data rows
data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        cell.alignment = data_alignment

# Specify the name of the merged file
merged_file_name = input("Enter the name of the merged file (without extension): ")
merged_file_path = file1_path.replace(file1_path.split("/")[-1], f"{merged_file_name}.xlsx")

# Save the workbook
wb.save(merged_file_path)

print(f"Merged data has been saved to: {merged_file_path}")
